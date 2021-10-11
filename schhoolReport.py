import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def schoolReport(xlsxFile):
    #open .xlxs
    if xlsxFile == None or xlsxFile == "":
        xlsxFile = "test.xlsx"
    wb = Workbook(xlsxFile)
    wb = openpyxl.load_workbook(xlsxFile)
    ws = wb.active
    #get all values from list
    grades = []
    names = []
    allGrades = {}

    #get names
    sheet = wb[wb.sheetnames[0]]
    for i in range(0,13):
        sourceValue = sheet.cell(row=10+i, column=1).value
        names.append(sourceValue)

    #get grades
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        if "9c" in sheetname:
            break
        grades = []

        for i in range(0,13):
            sourceValue = sheet.cell(row=10+i, column=2).value
            grades.append(sourceValue)
        allGrades[sheetname] = grades

    i = 0
    for name in names:
        print(name)

        for key in allGrades:
            print(key, allGrades[key][i] )
        i += 1





if __name__ == "__main__":
    xlsxFile = "/home/alex/schoolReport/Notenliste-11-10-2021_04-02-21.xlsx"
    schoolReport(xlsxFile)
