from schoolReport import getNamesGrades
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def testDictionary(xlsxFile):
    """ test add activity and schooltests"""
    names, allGrades = getNamesGrades.getNamesGrades(xlsxFile)
    if xlsxFile == None or xlsxFile == "":
        xlsxFile = "test.xlsx"
    wb = Workbook(xlsxFile)
    wb = openpyxl.load_workbook(xlsxFile)
    ws = wb.active

    def addGrade(gradeName, gradeColumn):
        """ adds Grade, first Argument is string of grade Name,
        second Argument is integer of column"""

        grades = []
        sheet = wb[wb.sheetnames[0]]
        for i in range(0,13):
            sourceValue = sheet.cell(row=10+i, column=gradeColumn).value
            grades.append(sourceValue)
        allGrades[wb.sheetnames[0]] [gradeName] = grades
        return allGrades

    xlsXCoordinates = [ ("Ø Mitarbeit", 6),
                        ("Ø Klausur", 3)]
    for i in range(len(xlsXCoordinates)):
        addGrade(xlsXCoordinates[i][0], xlsXCoordinates[i][1] )

    return names, allGrades

def gradesOfOnePupil(allGrades, pupilNumber):
    """print List of Grades of one pupil"""
    for key in allGrades:
        subject = key
        for key in allGrades[subject]:
            notenName = key
            note = allGrades[subject][notenName][pupilNumber]
            print(subject, notenName, note)




if __name__ == "__main__":
    xlsxFile = "/home/alex/schoolReport/Notenliste-11-10-2021_04-02-21.xlsx"
    names, allGrades = testDictionary(xlsxFile)

    #print List of Grades ordered by pupils
    pupilNumber = 0
    for name in names:
        print(name)
        gradesOfOnePupil(allGrades, pupilNumber)
        pupilNumber += 1
    print(allGrades)
