import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def getNamesGrades(xlsxFile):
    """ function to get from given xlsx File a List of names of pupil
    get a dictionary with subjects as keys and a List of Grades Gesamtnote """
    #open .xlxs
    if xlsxFile == None or xlsxFile == "":
        xlsxFile = "test.xlsx"
    wb = Workbook(xlsxFile)
    print("will intent to open the xlsx File")
    wb = openpyxl.load_workbook(xlsxFile)
    print("loaded")
    ws = wb.active
    #get all values from list
    grades = []
    names = []
    allGrades = {}

    #get names
    sheet = wb[wb.sheetnames[0]]
    for i in range(0,13):
        sourceValue = sheet.cell(row=10+i, column=1).value
        names.append(sourceValue)

    #get grades
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        if "9c" in sheetname:
            break

        #get Gesamtnote
        grades = []
        cNote = [["Gesamtnote",2]]
        for i in range(0,13):
            sourceValue = sheet.cell(row=10+i, column=cNote[0][1]).value
            grades.append(sourceValue)
        #allGrades[sheetname] = grades
        allGrades[sheetname] = {}
        allGrades[sheetname][cNote[0][0]] = grades


    return names, allGrades

if __name__ == "__main__":
    #xlsxFile = "/home/alex/schoolReport/Notenliste-11-10-2021_04-02-21.xlsx"
    xlsxFile = "/home/alex/schoolReport/Notenliste-18-10-2021_07-32-53.xlsx"
    wb = Workbook(xlsxFile)
    print("will intent to open the xlsx File")
    wb = openpyxl.load_workbook(xlsxFile)
    print("loaded")
    ws = wb.active



    """names, allGrades = getNamesGrades(xlsxFile)
    print(allGrades)
    i = 0
    for name in names:
        print(name)
        for key in allGrades:
            subject = key
            #print(subject, allGrades[subject] )
            for key in allGrades[subject]:
                print(subject, key, allGrades[subject][key][i])
            #print(key)
        i += 1"""
