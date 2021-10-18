import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def getParticipation(xlsxFile):
    """ function to get from given xlsx File a List of names of pupil
    get a dictionary with subjects as keys and a List of Grades Gesamtnote """
    #open .xlxs
    if xlsxFile == None or xlsxFile == "":
        xlsxFile = "test.xlsx"
    wb = Workbook(xlsxFile)
    print("will intent to open the xlsx File")
    wb = openpyxl.load_workbook(xlsxFile)
    print("loaded")
    ws = wb.active
    #get all values from list
    grades = []
    names = []
    fehlzeit = []
    allGrades = {}


    #get names
    sheet = wb[wb.sheetnames[0]]
    for i in range(0,13):
        sourceValue = sheet.cell(row=7+i, column=1).value
        names.append(sourceValue)

    FehlzeitenCo = [("Stu", 5),
                    ("Stu_ue", 8),
                    ("v", 9),]

    for zeit in FehlzeitenCo:
        #get Fehlzeiten
        sheet = wb[wb.sheetnames[0]]
        fehlzeit = []
        for i in range(0,13):
            sourceValue = sheet.cell(row=7+i, column=zeit[1]).value
            fehlzeit.append(sourceValue)
        print("fehlzeit: ", fehlzeit)
        allGrades[zeit[0]] = fehlzeit
        print(zeit[0])
    print(allGrades)

    return names, allGrades

if __name__ == "__main__":
    #xlsxFile = "/home/alex/schoolReport/Notenliste-11-10-2021_04-02-21.xlsx"
    #xlsxFile = "/home/alex/schoolReport/Notenliste-18-10-2021_07-32-53.xlsx"
    xlsxFile = "/home/alex/schoolReport/Unterrichtsstunden-18-10-2021_15-15-57.xlsx"
    wb = Workbook(xlsxFile)
    print("will intent to open the xlsx File")
    wb = openpyxl.load_workbook(xlsxFile)
    print("loaded")
    ws = wb.active
    getParticipation(xlsxFile)
